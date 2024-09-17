import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
from scipy import stats
import customtkinter as ctk  # Замена tkinter на customtkinter
from tkinter import filedialog, messagebox, scrolledtext, simpledialog, Listbox, MULTIPLE
from openpyxl import load_workbook

class CSVAnalyzer:
    def __init__(self, root):
        self.root = root
        self.root.title("CSV|Excel Analyzer")
        self.root.geometry("400x500")

        self.file_path = ""
        self.df = None

        self.create_widgets()

    def create_widgets(self):
        self.load_button = ctk.CTkButton(self.root, text="Загрузить CSV/Excel", command=self.load_file)
        self.load_button.pack(pady=10)

        self.view_button = ctk.CTkButton(self.root, text="Просмотреть данные", command=self.view_data)
        self.view_button.pack(pady=10)

        self.remove_duplicates_button = ctk.CTkButton(self.root, text="Удалить дубликаты", command=self.remove_duplicates)
        self.remove_duplicates_button.pack(pady=10)

        self.handle_missing_button = ctk.CTkButton(self.root, text="Обработать пропущенные значения", command=self.handle_missing)
        self.handle_missing_button.pack(pady=10)

        self.find_outliers_button = ctk.CTkButton(self.root, text="Найти выбросы", command=self.find_outliers)
        self.find_outliers_button.pack(pady=10)

        self.visualize_button = ctk.CTkButton(self.root, text="Визуализировать данные", command=self.select_visualization)
        self.visualize_button.pack(pady=10)

        self.stats_button = ctk.CTkButton(self.root, text="Статистический анализ", command=self.select_statistical_analysis)
        self.stats_button.pack(pady=10)

        self.save_button = ctk.CTkButton(self.root, text="Сохранить изменения", command=self.save_file)
        self.save_button.pack(pady=10)

        self.filter_button = ctk.CTkButton(self.root, text="Фильтровать строки", command=self.filter_data)
        self.filter_button.pack(pady=10)

        self.filter_columns_button = ctk.CTkButton(self.root, text="Фильтровать столбцы", command=self.filter_columns)
        self.filter_columns_button.pack(pady=10)

    def load_file(self):
        self.file_path = filedialog.askopenfilename(filetypes=[("CSV files", "*.csv"), ("Excel files", "*.xlsx;*.xls")])
        if self.file_path:
            if self.file_path.endswith('.csv'):
                self.df = pd.read_csv(self.file_path)
                messagebox.showinfo("Успех", "Файл успешно загружен!")
            elif self.file_path.endswith(('.xlsx', '.xls')):
                sheets = self.get_excel_sheets(self.file_path)
                if sheets:
                    sheet_name = simpledialog.askstring("Выбор листа", "Выберите лист:\n" + "\n".join(sheets))
                    if sheet_name:
                        self.df = pd.read_excel(self.file_path, sheet_name=sheet_name)
                        messagebox.showinfo("Успех", "Файл успешно загружен!")
                    else:
                        messagebox.showwarning("Ошибка", "Лист не выбран!")
                else:
                    messagebox.showwarning("Ошибка", "Не удалось получить список листов!")

    def get_excel_sheets(self, file_path):
        try:
            wb = load_workbook(file_path, read_only=True, keep_links=False)
            return wb.sheetnames
        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось открыть файл: {e}")
            return None

    def view_data(self):
        if self.df is not None:
            data_window = ctk.CTkToplevel(self.root)
            data_window.title("Просмотр данных")
            data_window.geometry("800x600")

            data_text = scrolledtext.ScrolledText(data_window, wrap='word', width=100, height=40)
            data_text.pack(pady=10)
            data_text.insert('end', self.df.to_string())
        else:
            messagebox.showwarning("Ошибка", "Сначала загрузите файл!")

    def save_file(self):
        if self.df is not None:
            file_path = filedialog.asksaveasfilename(defaultextension=".csv", filetypes=[("CSV files", "*.csv"), ("Excel files", "*.xlsx")])
            if file_path:
                if file_path.endswith('.csv'):
                    self.df.to_csv(file_path, index=False)
                elif file_path.endswith('.xlsx'):
                    self.df.to_excel(file_path, index=False)
                messagebox.showinfo("Успех", "Файл успешно сохранен!")
        else:
            messagebox.showwarning("Ошибка", "Сначала загрузите файл!")

    def remove_duplicates(self):
        if self.df is not None:
            duplicates = self.df[self.df.duplicated()]
            if not duplicates.empty:
                duplicate_window = ctk.CTkToplevel(self.root)
                duplicate_window.title("Дубликаты")
                duplicate_window.geometry("800x600")

                duplicate_text = scrolledtext.ScrolledText(duplicate_window, wrap='word', width=100, height=40)
                duplicate_text.pack(pady=10)
                duplicate_text.insert('end', duplicates.to_string())

                def on_ok():
                    duplicate_window.destroy()
                    self.df.drop_duplicates(inplace=True)
                    messagebox.showinfo("Успех", "Дубликаты удалены!")

                ok_button = ctk.CTkButton(duplicate_window, text="OK", command=on_ok)
                ok_button.pack(pady=10)
            else:
                messagebox.showinfo("Успех", "Дубликаты не найдены!")
        else:
            messagebox.showwarning("Ошибка", "Сначала загрузите файл!")

    def handle_missing(self):
        if self.df is not None:
            choice = simpledialog.askstring("Обработка пропущенных данных", "Выберите метод обработки:\n1. Удалить строки\n2. Заполнить средними\n3. Заполнить медианой\n4. Заполнить самым частым значением")
            if choice == '1':
                self.df.dropna(inplace=True)
            elif choice == '2':
                self.df.fillna(self.df.mean(), inplace=True)
            elif choice == '3':
                self.df.fillna(self.df.median(), inplace=True)
            elif choice == '4':
                self.df.fillna(self.df.mode().iloc[0], inplace=True)
            messagebox.showinfo("Успех", "Пропущенные значения обработаны!")
        else:
            messagebox.showwarning("Ошибка", "Сначала загрузите файл!")

    def find_outliers(self):
        if self.df is not None:
            z_scores = np.abs(stats.zscore(self.df.select_dtypes(include=[np.number])))
            outliers = (z_scores > 3)
            messagebox.showinfo("Выбросы", f"Найдено {outliers.sum().sum()} выбросов.")
        else:
            messagebox.showwarning("Ошибка", "Сначала загрузите файл!")

    def select_visualization(self):
        if self.df is not None:
            vis_types = [
                "Столбчатые диаграммы", "Гистограмма", "Линейный график", "Многоуровневая диаграмма",
                "Точечный график", "Тепловая карта", "Диаграмма воронки", "Круговая диаграмма",
                "Диаграмма размаха", "Ядерная оценка плотности", "Кумулятивное распределение",
                "График квантиль-квантиль", "Скрипичная диаграмма", "Ridgeline (групповая хребтовая диаграмма)"
            ]
            vis_type = simpledialog.askstring("Выбор типа визуализации", "Выберите тип визуализации:\n" + "\n".join(f"{i+1}. {t}" for i, t in enumerate(vis_types)))
            if vis_type:
                vis_type = int(vis_type) - 1
                if 0 <= vis_type < len(vis_types):
                    self.select_columns(vis_types[vis_type])
                else:
                    messagebox.showwarning("Ошибка", "Неверный выбор!")
        else:
            messagebox.showwarning("Ошибка", "Сначала загрузите файл!")

    def select_columns(self, vis_type):
        if self.df is not None:
            columns = self.df.columns.tolist()
            column_window = ctk.CTkToplevel(self.root)
            column_window.title("Выбор столбцов")
            column_window.geometry("300x300")

            listbox = Listbox(column_window, selectmode=MULTIPLE)
            listbox.pack(fill='both', expand=True)

            for col in columns:
                listbox.insert('end', col)

            def on_ok():
                selected_columns = [listbox.get(i) for i in listbox.curselection()]
                column_window.destroy()
                self.visualize_data(vis_type, selected_columns)

            ok_button = ctk.CTkButton(column_window, text="OK", command=on_ok)
            ok_button.pack(pady=10)
        else:
            messagebox.showwarning("Ошибка", "Сначала загрузите файл!")

    def visualize_data(self, vis_type, selected_columns):
        if self.df is not None:
            plt.figure(figsize=(10, 6))
            if vis_type == "Столбчатые диаграммы":
                self.df[selected_columns].plot(kind='bar')
            elif vis_type == "Гистограмма":
                self.df[selected_columns].hist()
            elif vis_type == "Линейный график":
                self.df[selected_columns].plot()
            elif vis_type == "Многоуровневая диаграмма":
                sns.violinplot(data=self.df[selected_columns])
            elif vis_type == "Точечный график":
                sns.scatterplot(data=self.df[selected_columns])
            elif vis_type == "Тепловая карта":
                sns.heatmap(self.df[selected_columns].corr(), annot=True, cmap='coolwarm')
            elif vis_type == "Диаграмма воронки":
                sns.barplot(x=self.df[selected_columns[0]], y=self.df[selected_columns[1]])
            elif vis_type == "Круговая диаграмма":
                self.df[selected_columns].plot(kind='pie', subplots=True)
            elif vis_type == "Диаграмма размаха":
                sns.boxplot(data=self.df[selected_columns])
            elif vis_type == "Ядерная оценка плотности":
                sns.kdeplot(data=self.df[selected_columns])
            elif vis_type == "Кумулятивное распределение":
                sns.ecdfplot(data=self.df[selected_columns])
            elif vis_type == "График квантиль-квантиль":
                stats.probplot(self.df[selected_columns[0]], plot=plt)
            elif vis_type == "Скрипичная диаграмма":
                sns.violinplot(data=self.df[selected_columns])
            elif vis_type == "Ridgeline (групповая хребтовая диаграмма)":
                sns.kdeplot(data=self.df[selected_columns], shade=True)
            plt.show()
        else:
            messagebox.showwarning("Ошибка", "Сначала загрузите файл!")

    def select_statistical_analysis(self):
        if self.df is not None:
            analysis_types = [
                "Корреляционный анализ", "Регрессия", "Дисперсия", "Стандартное отклонение",
                "Диапазон", "Меры центральной тенденции"
            ]
            analysis_type = simpledialog.askstring("Выбор типа статистического анализа", "Выберите тип анализа:\n" + "\n".join(f"{i+1}. {t}" for i, t in enumerate(analysis_types)))
            if analysis_type:
                analysis_type = int(analysis_type) - 1
                if 0 <= analysis_type < len(analysis_types):
                    self.select_columns_for_analysis(analysis_types[analysis_type])
                else:
                    messagebox.showwarning("Ошибка", "Неверный выбор!")
        else:
            messagebox.showwarning("Ошибка", "Сначала загрузите файл!")

    def select_columns_for_analysis(self, analysis_type):
        if self.df is not None:
            columns = self.df.columns.tolist()
            column_window = ctk.CTkToplevel(self.root)
            column_window.title("Выбор столбцов")
            column_window.geometry("300x300")

            listbox = Listbox(column_window, selectmode=MULTIPLE)
            listbox.pack(fill='both', expand=True)

            for col in columns:
                listbox.insert('end', col)

            def on_ok():
                selected_columns = [listbox.get(i) for i in listbox.curselection()]
                column_window.destroy()
                self.perform_statistical_analysis(analysis_type, selected_columns)

            ok_button = ctk.CTkButton(column_window, text="OK", command=on_ok)
            ok_button.pack(pady=10)
        else:
            messagebox.showwarning("Ошибка", "Сначала загрузите файл!")

    def perform_statistical_analysis(self, analysis_type, selected_columns):
        if self.df is not None:
            if analysis_type == "Корреляционный анализ":
                result = self.df[selected_columns].corr()
            elif analysis_type == "Регрессия":
                result = "Регрессия: Не реализовано"
            elif analysis_type == "Дисперсия":
                result = self.df[selected_columns].var()
            elif analysis_type == "Стандартное отклонение":
                result = self.df[selected_columns].std()
            elif analysis_type == "Диапазон":
                result = self.df[selected_columns].max() - self.df[selected_columns].min()
            elif analysis_type == "Меры центральной тенденции":
                mean = self.df[selected_columns].mean()
                median = self.df[selected_columns].median()
                mode = self.df[selected_columns].mode().iloc[0]
                result = f"Среднее арифметическое: {mean}\nМедиана: {median}\nМода: {mode}"

            messagebox.showinfo("Статистический анализ", result.to_string() if isinstance(result, pd.Series) else result)
        else:
            messagebox.showwarning("Ошибка", "Сначала загрузите файл!")

    def filter_data(self):
        if self.df is not None:
            filter_window = ctk.CTkToplevel(self.root)
            filter_window.title("Фильтрация строк")
            filter_window.geometry("400x400")

            ctk.CTkLabel(filter_window, text="Выберите столбец:").pack(pady=5)
            column_var = ctk.StringVar(filter_window)
            column_var.set(self.df.columns[0])
            column_menu = ctk.CTkOptionMenu(filter_window, variable=column_var, values=list(self.df.columns))
            column_menu.pack(pady=5)

            ctk.CTkLabel(filter_window, text="Введите значение для фильтрации:").pack(pady=5)
            value_entry = ctk.CTkEntry(filter_window)
            value_entry.pack(pady=5)

            ctk.CTkLabel(filter_window, text="Введите диапазон строк (например, 1-10):").pack(pady=5)
            range_entry = ctk.CTkEntry(filter_window)
            range_entry.pack(pady=5)

            def apply_filter():
                column = column_var.get()
                value = value_entry.get()
                range_str = range_entry.get()

                if value:
                    self.df = self.df[self.df[column] == value]

                if range_str:
                    try:
                        start, end = map(int, range_str.split('-'))
                        self.df = self.df.iloc[start-1:end]
                    except ValueError:
                        messagebox.showwarning("Ошибка", "Неверный формат диапазона строк!")

                messagebox.showinfo("Успех", "Фильтрация применена!")
                filter_window.destroy()

            apply_button = ctk.CTkButton(filter_window, text="Применить фильтр", command=apply_filter)
            apply_button.pack(pady=10)
        else:
            messagebox.showwarning("Ошибка", "Сначала загрузите файл!")

    def filter_columns(self):
        if self.df is not None:
            filter_window = ctk.CTkToplevel(self.root)
            filter_window.title("Фильтрация столбцов")
            filter_window.geometry("300x300")

            ctk.CTkLabel(filter_window, text="Выберите столбцы для исключения:").pack(pady=5)
            listbox = Listbox(filter_window, selectmode=MULTIPLE)
            listbox.pack(fill='both', expand=True)

            for col in self.df.columns:
                listbox.insert('end', col)

            def apply_filter():
                selected_columns = [listbox.get(i) for i in listbox.curselection()]
                self.df = self.df.drop(columns=selected_columns)
                messagebox.showinfo("Успех", "Фильтрация столбцов применена!")
                filter_window.destroy()

            apply_button = ctk.CTkButton(filter_window, text="Применить фильтр", command=apply_filter)
            apply_button.pack(pady=10)
        else:
            messagebox.showwarning("Ошибка", "Сначала загрузите файл!")

if __name__ == "__main__":
    ctk.set_appearance_mode("dark")  # Опционально: режим темной темы
    ctk.set_default_color_theme("blue")  # Опционально: выбор цветовой темы

    root = ctk.CTk()  # Используем CTk вместо Tk
    app = CSVAnalyzer(root)
    root.mainloop()
